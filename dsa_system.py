import openpyxl
import os
import json
import argparse
import urllib.request
from datetime import datetime, timedelta
from collections import defaultdict

EXCEL_PATH = "C:/Users/bhuva/OneDrive/Desktop/DSA/DSA.xlsx"
TRACKER_DIR = "C:/Users/bhuva/OneDrive/Desktop/DSA/DSA_Tracker"
DB_PATH = os.path.join(TRACKER_DIR, "revision_state.json")
DAILY_PLAN_PATH = os.path.join(TRACKER_DIR, "Daily_Plan.md")

# LeetSync sometimes syncs folders with wrong problem numbers.
# This map corrects known mismatches: { wrong_folder_id: correct_leetcode_id }
LEETSYNC_ID_CORRECTIONS = {
    "1056": "1011",  # Capacity To Ship Packages Within D Days
    "1078": "1021",  # Remove Outermost Parentheses
    "1128": "1047",  # Remove All Adjacent Duplicates In String
    "745": "744",    # Find Smallest Letter Greater Than Target
    "874": "844",    # Backspace String Compare
    "882": "852",    # Peak Index in a Mountain Array
    "907": "875",    # Koko Eating Bananas
}

def load_db():
    if os.path.exists(DB_PATH):
        with open(DB_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"problems": {}, "patterns": {}, "mistakes": []}

def save_db(db):
    os.makedirs(TRACKER_DIR, exist_ok=True)
    with open(DB_PATH, "w", encoding="utf-8") as f:
        json.dump(db, f, indent=4)

def calculate_next_review(ease, interval, performance):
    """
    SuperMemo-2 based logic.
    Performance: 3 = Easy, 2 = Medium, 1 = Hard/Hint, 0 = Failed
    """
    if performance < 1:
        return 1, max(1.3, ease - 0.2)
    
    if interval == 0:
        next_interval = 1
    elif interval == 1:
        next_interval = 3
    else:
        next_interval = int(interval * ease)
        
    new_ease = ease + (0.1 - (3 - performance) * (0.08 + (3 - performance) * 0.02))
    return next_interval, max(1.3, new_ease)

def fetch_real_leetcode_id(slug):
    try:
        url = 'https://leetcode.com/graphql'
        data = json.dumps({
            'query': 'query questionTitle($titleSlug: String!) { question(titleSlug: $titleSlug) { questionFrontendId title } }',
            'variables': {'titleSlug': slug}
        }).encode('utf-8')
        req = urllib.request.Request(url, data=data, headers={
            'Content-Type': 'application/json',
            'User-Agent': 'Mozilla/5.0',
            'Referer': 'https://leetcode.com/'
        })
        resp = urllib.request.urlopen(req, timeout=10)
        resp_data = json.loads(resp.read().decode('utf-8'))
        return resp_data['data']['question']['questionFrontendId']
    except Exception as e:
        print(f"Failed to fetch ID for {slug}: {e}")
        return None

def sync_excel():
    db = load_db()
    
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"Error loading {EXCEL_PATH}: {e}")
        return

    added = 0
    for row in range(5, ws.max_row + 1):
        pid_val = ws.cell(row=row, column=1).value
        status = ws.cell(row=row, column=11).value
        
        if status == '☑' and pid_val is not None:
            try:
                pid = str(int(pid_val))
                if pid not in db["problems"]:
                    db["problems"][pid] = {
                        "name": ws.cell(row=row, column=2).value or "Unknown",
                        "difficulty": ws.cell(row=row, column=9).value or "Unknown",
                        "pattern": ws.cell(row=row, column=8).value or "General",
                        "ease": 2.5,
                        "interval": 0,
                        "next_review": datetime.now().strftime("%Y-%m-%d"),
                        "history": [],
                        "state": "New"
                    }
                    added += 1
            except ValueError:
                pass
                
    print(f"Synced from Excel. {added} new problems added to revision state.")
    
    # Sync from LeetSync folders
    leetsync_added = 0
    workspace = "C:/Users/bhuva/OneDrive/Desktop/DSA"
    for item in os.listdir(workspace):
        item_path = os.path.join(workspace, item)
        if os.path.isdir(item_path) and item[0].isdigit() and "-" in item:
            try:
                parts = item.split("-", 1)
                folder_id_str = parts[0]
                slug = parts[1]
                pid = str(int(folder_id_str))
                
                # Apply known LeetSync ID corrections
                pid = LEETSYNC_ID_CORRECTIONS.get(pid, pid)
                
                if pid not in db["problems"]:
                    # Auto-Fixer: Verify with LeetCode API
                    real_id = fetch_real_leetcode_id(slug)
                    if real_id and real_id != pid:
                        print(f"Auto-Fixer detected mismatch: Folder ID {pid}, Real ID {real_id}. Fixing...")
                        # 1. Add to LEETSYNC_ID_CORRECTIONS in memory
                        LEETSYNC_ID_CORRECTIONS[folder_id_str] = real_id
                        
                        # 2. Add to DSA_Tracker/leetsync_id_corrections.json
                        corrections_file = os.path.join(TRACKER_DIR, "leetsync_id_corrections.json")
                        if os.path.exists(corrections_file):
                            try:
                                with open(corrections_file, "r", encoding="utf-8") as cf:
                                    corr_db = json.load(cf)
                                corr_db[folder_id_str] = {
                                    "correct_id": real_id,
                                    "title": slug.replace("-", " ").title(),
                                    "folder": f"{real_id}-{slug}",
                                    "fixed_date": datetime.now().strftime("%Y-%m-%d")
                                }
                                with open(corrections_file, "w", encoding="utf-8") as cf:
                                    json.dump(corr_db, cf, indent=4)
                            except Exception as ce:
                                print(f"Error logging correction: {ce}")
                        
                        # 3. Rename folder
                        new_folder_name = f"{real_id}-{slug}"
                        new_item_path = os.path.join(workspace, new_folder_name)
                        try:
                            os.rename(item_path, new_item_path)
                            item_path = new_item_path
                            item = new_folder_name
                            pid = real_id
                            print(f"Successfully renamed folder to {new_folder_name}")
                        except Exception as re:
                            print(f"Failed to rename folder {item}: {re}")

                    if pid not in db["problems"]:
                        # Try to extract difficulty from README
                        difficulty = "Unknown"
                        readme_path = os.path.join(item_path, "README.md")
                        if os.path.exists(readme_path):
                            with open(readme_path, "r", encoding="utf-8") as f:
                                content = f.read()
                                if "Difficulty-Easy" in content: difficulty = "Easy"
                                elif "Difficulty-Medium" in content: difficulty = "Medium"
                                elif "Difficulty-Hard" in content: difficulty = "Hard"
                                
                        name = slug.replace("-", " ").title()
                        
                        db["problems"][pid] = {
                            "name": name,
                            "difficulty": difficulty,
                            "pattern": "General (LeetSync)",
                            "ease": 2.5,
                            "interval": 0,
                            "next_review": datetime.now().strftime("%Y-%m-%d"),
                            "history": [],
                            "state": "New"
                        }
                        leetsync_added += 1
            except Exception as e:
                print(f"Error parsing LeetSync folder {item}: {e}")
                
    if leetsync_added > 0:
        save_db(db)
        print(f"Synced from LeetSync. {leetsync_added} new problems added to revision state.")
        
    print(f"Total problems in system: {len(db['problems'])}")

def generate_daily_plan():
    db = load_db()
    today = datetime.now().date()
    
    due = []
    weak = []
    for pid, data in db["problems"].items():
        review_date = datetime.strptime(data["next_review"], "%Y-%m-%d").date()
        if review_date <= today:
            due.append((pid, data))
        if data["state"] == "Fragile":
            weak.append((pid, data))
            
    # Sort due by interval (shortest interval first)
    due.sort(key=lambda x: x[1]["interval"])
    
    with open(DAILY_PLAN_PATH, "w", encoding="utf-8") as f:
        f.write(f"# 📅 DSA Daily Plan — {today.strftime('%b %d, %Y')}\n\n")
        f.write("## 🆕 NEW (Goal: 1-2 Problems)\n")
        f.write("- Solve on LeetCode, then log in `DSA.xlsx` and run `python dsa_system.py sync`.\n\n")
        
        f.write("## 🧠 RETRIEVAL REVIEW (Due Today)\n")
        if not due:
            f.write("*Nothing due today! Relax or solve new problems.*\n")
        else:
            for pid, p in due[:5]:  # Cap at 5 for sustainability
                f.write(f"- [ ] **{p['name']}** (ID: {pid}) — Pattern: *{p['pattern']}*\n")
        f.write("\n")
        
        f.write("## ⚠️ WEAK PATTERNS (Needs Reinforcement)\n")
        if not weak:
            f.write("*No fragile patterns detected.*\n")
        else:
            for pid, p in weak[:2]:
                f.write(f"- [ ] **{p['name']}** (ID: {pid}) — Pattern: *{p['pattern']}*\n")
        f.write("\n")
        
        f.write("---\n")
        f.write("**How to log a review:**\n")
        f.write("Run: `python dsa_system.py review [ID]`\n")
        
    print(f"Daily Plan generated at: {DAILY_PLAN_PATH}")

def run_review(pid):
    db = load_db()
    pid = str(pid)
    
    if pid not in db["problems"]:
        print(f"Error: Problem ID {pid} not found in system.")
        return
        
    p = db["problems"][pid]
    print(f"\nReviewing: {p['name']} (Pattern: {p['pattern']})")
    
    print("\nHow was the recall?")
    print("3 - Easy (Immediate pattern recognition & implementation)")
    print("2 - Medium (Correct, but slow or uncertain)")
    print("1 - Hard (Needed a hint)")
    print("0 - Failed (Wrong pattern or completely stuck)")
    
    while True:
        try:
            perf = int(input("Select score (0-3): "))
            if perf in [0, 1, 2, 3]:
                break
        except:
            pass
            
    mistake = input("Main mistake or takeaway (leave blank if none): ")
    
    interval, ease = calculate_next_review(p["ease"], p["interval"], perf)
    next_date = datetime.now() + timedelta(days=interval)
    
    p["ease"] = ease
    p["interval"] = interval
    p["next_review"] = next_date.strftime("%Y-%m-%d")
    
    if perf == 0:
        p["state"] = "Fragile"
    elif interval > 21:
        p["state"] = "Mastered"
    elif interval > 7:
        p["state"] = "Stable"
    else:
        p["state"] = "Learning"
        
    if mistake:
        db["mistakes"].append({
            "date": datetime.now().strftime("%Y-%m-%d"),
            "pid": pid,
            "pattern": p["pattern"],
            "mistake": mistake
        })
        
    save_db(db)
    print(f"\nReview saved! Next review in {interval} days ({p['next_review']}). State: {p['state']}.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="DSA Adaptive Revision System")
    parser.add_argument("command", choices=["sync", "daily", "review"], help="Command to run")
    parser.add_argument("pid", nargs="?", help="Problem ID for review")
    
    args = parser.parse_args()
    
    if args.command == "sync":
        sync_excel()
    elif args.command == "daily":
        generate_daily_plan()
    elif args.command == "review":
        if not args.pid:
            print("Please provide a Problem ID: python dsa_system.py review [ID]")
        else:
            run_review(args.pid)
