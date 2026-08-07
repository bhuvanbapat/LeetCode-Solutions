import openpyxl
import os
from collections import defaultdict

file_path = "C:/Users/bhuva/OneDrive/Desktop/DSA/DSA.xlsx"
tracker_dir = "C:/Users/bhuva/OneDrive/Desktop/DSA/DSA_Tracker"
os.makedirs(tracker_dir, exist_ok=True)

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
except Exception as e:
    print(f"Error loading workbook: {e}")
    exit(1)

solved_problems = []
difficulty_counts = {"Easy": 0, "Medium": 0, "Hard": 0}
topic_counts = defaultdict(int)

# Extract data
for row in range(5, ws.max_row + 1):
    val_a = ws.cell(row=row, column=1).value
    status = ws.cell(row=row, column=11).value
    if status == '☑' and val_a is not None:
        try:
            problem_id = int(val_a)
            problem_name = ws.cell(row=row, column=2).value or "Unknown"
            difficulty = ws.cell(row=row, column=9).value or "Unknown"
            pattern = ws.cell(row=row, column=8).value or "General"
            
            solved_problems.append({
                "id": problem_id,
                "name": problem_name,
                "difficulty": difficulty,
                "pattern": pattern
            })
            
            if difficulty in difficulty_counts:
                difficulty_counts[difficulty] += 1
            else:
                if difficulty.startswith("Med"):
                    difficulty_counts["Medium"] += 1
                else:
                    pass # Custom or unknown
            
            topic_counts[pattern] += 1
            
        except ValueError:
            pass

# Create Dashboard.md
dashboard_path = os.path.join(tracker_dir, "Dashboard.md")
with open(dashboard_path, "w", encoding="utf-8") as f:
    f.write("# 🚀 DSA Progress Dashboard\n\n")
    f.write(f"**Total Problems Solved:** {len(solved_problems)}\n\n")
    f.write("## 📊 Difficulty Breakdown\n")
    f.write(f"- 🟢 **Easy:** {difficulty_counts['Easy']}\n")
    f.write(f"- 🟡 **Medium:** {difficulty_counts['Medium']}\n")
    f.write(f"- 🔴 **Hard:** {difficulty_counts['Hard']}\n\n")
    f.write("## 🧩 Top Patterns / Topics\n")
    sorted_topics = sorted(topic_counts.items(), key=lambda x: x[1], reverse=True)
    for topic, count in sorted_topics:
        f.write(f"- **{topic}:** {count} problems\n")

# Create Daily_Log.md
log_path = os.path.join(tracker_dir, "Daily_Log.md")
with open(log_path, "w", encoding="utf-8") as f:
    f.write("# 📅 DSA Daily Log\n\n")
    f.write("*Track your daily problem-solving journey below. Add notes and takeaways for each problem!* \n\n")
    f.write("## Initial Import\n")
    f.write(f"*The following {len(solved_problems)} problems were imported from your initial progress.* \n\n")
    f.write("| Problem ID | Name | Difficulty | Pattern | Notes/Takeaways |\n")
    f.write("|---|---|---|---|---|\n")
    for p in solved_problems:
        diff_icon = "🟢" if "Easy" in p['difficulty'] else "🟡" if "Med" in p['difficulty'] else "🔴" if "Hard" in p['difficulty'] else ""
        f.write(f"| {p['id']} | **{p['name']}** | {diff_icon} {p['difficulty']} | {p['pattern']} | *Add notes here* |\n")

# Create AI_Update_Prompt.md
prompt_path = os.path.join(tracker_dir, "AI_Update_Prompt.md")
with open(prompt_path, "w", encoding="utf-8") as f:
    f.write("# 🤖 AI Update Prompt\n\n")
    f.write("Copy and paste the following template into the chat whenever you finish new problems. This ensures I update everything accurately and without hallucinating.\n\n")
    f.write("---\n\n")
    f.write("### Template to Copy:\n\n")
    f.write("```markdown\n")
    f.write("Hey Antigravity, I just completed some new DSA problems! Please update my Excel sheet and my Markdown tracking system (`DSA_Tracker/Dashboard.md` and `DSA_Tracker/Daily_Log.md`). \n\n")
    f.write("Here are the new problems I solved:\n")
    f.write("- [Problem Number 1]\n")
    f.write("- [Problem Number 2]\n\n")
    f.write("**Instructions for you:**\n")
    f.write("1. Mark these problem numbers as '☑' in `C:/Users/bhuva/OneDrive/Desktop/DSA/DSA.xlsx`.\n")
    f.write("2. Read the updated Excel file using Python to recalculate my stats.\n")
    f.write("3. Update `DSA_Tracker/Dashboard.md` with the new stats (Total, Difficulty, Patterns).\n")
    f.write("4. Append these newly solved problems to `DSA_Tracker/Daily_Log.md` under a new header `## Date: [Today's Date]`, including the Problem ID, Name, Difficulty, and Pattern.\n")
    f.write("```\n")

print("Tracking system generated successfully.")
